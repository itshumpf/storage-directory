"""
Public Storage Location Scraper
Scrapes all US locations from publicstorage.com sitemaps
Outputs: public_storage_locations.xlsx
"""

import re
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
BASE_URL = "https://www.publicstorage.com"
DELAY = 0.5  # seconds between requests — be polite


def get_state_urls():
    """Fetch list of all state sitemap URLs."""
    url = f"{BASE_URL}/site-map-states"
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    state_urls = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "site-map-states-" in href:
            full = href if href.startswith("http") else BASE_URL + href
            state_urls.append(full)

    return list(set(state_urls))


def get_facility_urls(state_url):
    """Fetch all facility page URLs from a state sitemap page."""
    resp = requests.get(state_url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    facility_urls = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        # Facility pages match pattern: /self-storage-xx-cityname/NNNNN.html
        if re.search(r"/self-storage-[a-z]+-[a-z0-9-]+/\d{4,6}\.html$", href):
            full = href if href.startswith("http") else BASE_URL + href
            facility_urls.append((full, a.get_text(strip=True)))

    return facility_urls


def parse_facility_url(url):
    """Extract store number, state, and city from a facility URL."""
    # e.g. https://www.publicstorage.com/self-storage-ks-shawnee/77942.html
    match = re.search(
        r"/self-storage-([a-z]{2})-([a-z0-9-]+)/(\d{4,6})\.html$", url
    )
    if match:
        state = match.group(1).upper()
        city = match.group(2).replace("-", " ").title()
        store_num = match.group(3)
        return store_num, state, city
    return None, None, None


def get_facility_address(url):
    """Fetch the address from an individual facility page."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try common address selectors
        for selector in [
            '[itemprop="streetAddress"]',
            ".facility-address",
            ".property-address",
            '[class*="address"]',
        ]:
            el = soup.select_one(selector)
            if el:
                return el.get_text(strip=True)

        # Fallback: look for schema.org address in JSON-LD
        for script in soup.find_all("script", type="application/ld+json"):
            import json
            try:
                data = json.loads(script.string)
                if isinstance(data, dict) and "address" in data:
                    addr = data["address"]
                    if isinstance(addr, dict):
                        return addr.get("streetAddress", "")
            except Exception:
                pass

    except Exception as e:
        print(f"  Warning: could not fetch address from {url}: {e}")

    return ""


def save_to_xlsx(locations, output_path="public_storage_locations.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Locations"

    headers = ["Store Number", "Address", "City", "State", "URL"]
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, loc in enumerate(locations, 2):
        ws.cell(row=row_idx, column=1, value=loc["store_number"])
        ws.cell(row=row_idx, column=2, value=loc["address"])
        ws.cell(row=row_idx, column=3, value=loc["city"])
        ws.cell(row=row_idx, column=4, value=loc["state"])
        ws.cell(row=row_idx, column=5, value=loc["url"])

    col_widths = [14, 40, 24, 8, 70]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add total count
    summary_row = len(locations) + 3
    ws.cell(row=summary_row, column=1, value="Total Locations:")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, name="Arial")
    ws.cell(row=summary_row, column=2, value=f'=COUNTA(A2:A{len(locations)+1})')

    wb.save(output_path)
    print(f"\nSaved {len(locations)} locations to {output_path}")


def main():
    print("Step 1: Fetching state sitemap URLs...")
    state_urls = get_state_urls()
    print(f"  Found {len(state_urls)} states")

    all_facilities = []

    print("\nStep 2: Fetching facility URLs from each state...")
    for i, state_url in enumerate(sorted(state_urls)):
        state_name = state_url.split("site-map-states-")[-1].replace("-", " ").title()
        print(f"  [{i+1}/{len(state_urls)}] {state_name}...", end=" ", flush=True)

        try:
            facilities = get_facility_urls(state_url)
            print(f"{len(facilities)} facilities")
            all_facilities.extend(facilities)
        except Exception as e:
            print(f"ERROR: {e}")

        time.sleep(DELAY)

    print(f"\nTotal facility URLs found: {len(all_facilities)}")

    # Deduplicate
    seen = set()
    unique_facilities = []
    for url, label in all_facilities:
        if url not in seen:
            seen.add(url)
            unique_facilities.append((url, label))

    print(f"Unique facilities: {len(unique_facilities)}")

    print("\nStep 3: Parsing store data from URLs...")
    locations = []
    for url, label in unique_facilities:
        store_num, state, city = parse_facility_url(url)
        if store_num:
            # Try to extract address from the link label first
            # Label is often "Self Storage Near [ADDRESS] in [City, ST]"
            address = ""
            addr_match = re.search(r"Self Storage Near (.+?) in ", label, re.IGNORECASE)
            if addr_match:
                address = addr_match.group(1).strip()

            locations.append({
                "store_number": store_num,
                "address": address,
                "city": city,
                "state": state,
                "url": url,
            })

    print(f"Parsed {len(locations)} locations from URLs")

    # Optional: fetch full addresses from facility pages
    # This is slower (one request per store) — comment out if you just want the URL data
    FETCH_ADDRESSES = False  # Set to True to fetch individual pages for full addresses
    if FETCH_ADDRESSES:
        print("\nStep 4 (optional): Fetching full addresses from facility pages...")
        for i, loc in enumerate(locations):
            if not loc["address"]:
                print(f"  [{i+1}/{len(locations)}] Fetching {loc['store_number']}...")
                loc["address"] = get_facility_address(loc["url"])
                time.sleep(DELAY)

    print("\nStep 4: Saving to Excel...")
    save_to_xlsx(locations)


if __name__ == "__main__":
    main()
