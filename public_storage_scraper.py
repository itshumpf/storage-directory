# ============================================================
# PUBLIC STORAGE SCRAPER — Run in Google Colab
# Uses Public Storage's internal API (no scraping blocks!)
# Outputs: locations.json + public_storage_locations.xlsx
# ============================================================

# Step 1: Install dependencies
# !pip install requests beautifulsoup4 openpyxl

import requests
import json
import re
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.colab import files  # for downloading the file at the end

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.publicstorage.com/self-storage-search",
}

BASE = "https://www.publicstorage.com"
PRICING_API = BASE + "/on/demandware.store/Sites-publicstorage-Site/default/AP-GetSoostonePromo?sites={}"
SITEMAP_STATE_INDEX = BASE + "/site-map-states"
BATCH_SIZE = 20   # store IDs per API call
DELAY = 0.4       # seconds between requests


# ── Step 1: Get all store IDs from sitemaps ──────────────────

def get_state_sitemap_urls():
    resp = requests.get(SITEMAP_STATE_INDEX, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    urls = []
    for a in soup.find_all("a", href=True):
        if "site-map-states-" in a["href"]:
            href = a["href"]
            urls.append(href if href.startswith("http") else BASE + href)
    return list(set(urls))


def get_stores_from_state(state_url):
    """Returns list of dicts with store_id, address, city, state, url"""
    stores = []
    try:
        resp = requests.get(state_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            m = re.search(r"/self-storage-([a-z]{2})-([a-z0-9-]+)/(\d{4,6})\.html$", href)
            if m:
                state_code = m.group(1).upper()
                city = m.group(2).replace("-", " ").title()
                store_id = m.group(3)
                label = a.get_text(strip=True)
                # Extract street address from label like "Self Storage Near 123 Main St in City, ST"
                addr_m = re.search(r"Self Storage Near (.+?) in ", label, re.IGNORECASE)
                address = addr_m.group(1).strip() if addr_m else ""
                full_url = href if href.startswith("http") else BASE + href
                stores.append({
                    "store_id": store_id,
                    "address": address,
                    "city": city,
                    "state": state_code,
                    "url": full_url,
                    "units": []
                })
    except Exception as e:
        print(f"  Warning: {state_url} — {e}")
    return stores


# ── Step 2: Fetch pricing from API in batches ─────────────────

def fetch_pricing_batch(store_ids):
    """Fetch unit pricing for a batch of store IDs. Returns dict of store_id → units list."""
    joined = "%2C".join(store_ids)
    url = PRICING_API.format(joined)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        results = {}

        # Response structure: { promoInfoArr: [ {storeID, info: [{name, saleprice, availability, count, promotionName}]} ] }
        promo_arr = data.get("promoInfoArr", [])
        for store_data in promo_arr:
            sid = str(store_data.get("storeID", ""))
            units = []
            for unit in store_data.get("info", []):
                units.append({
                    "size": unit.get("name", ""),
                    "price": unit.get("saleprice", None),
                    "available": unit.get("availability", False),
                    "count": unit.get("count", 0),
                    "promo": unit.get("promotionName", "") or "",
                    "promo2": unit.get("promotionName2", "") or "",
                })
            results[sid] = units
        return results
    except Exception as e:
        print(f"  Pricing API error for batch: {e}")
        return {}


# ── Step 3: Save to Excel ─────────────────────────────────────

def save_xlsx(stores, filename="public_storage_locations.xlsx"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Locations ──
    ws1 = wb.active
    ws1.title = "Locations"

    hdr_fill = PatternFill("solid", start_color="1F3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="EEF2F7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    headers = ["Store #", "Address", "City", "State", "Cheapest Unit", "Lowest Price", "Units Available", "URL"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.row_dimensions[1].height = 22

    for i, store in enumerate(stores, 2):
        units = store.get("units", [])
        avail = [u for u in units if u["available"] and u["price"]]
        cheapest = min(avail, key=lambda u: u["price"]) if avail else None

        row_fill = alt_fill if i % 2 == 0 else None
        vals = [
            store["store_id"],
            store["address"],
            store["city"],
            store["state"],
            cheapest["size"] if cheapest else "",
            f"${cheapest['price']}/mo" if cheapest else "",
            sum(u["count"] for u in avail),
            store["url"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            c.border = border
            if row_fill:
                c.fill = row_fill

    col_widths = [10, 35, 22, 8, 14, 14, 16, 65]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = w
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Unit Pricing Detail ──
    ws2 = wb.create_sheet("Unit Pricing")
    p_headers = ["Store #", "City", "State", "Unit Size", "Sale Price", "Available", "Units Left", "Promotion"]
    for col, h in enumerate(p_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    row = 2
    for store in stores:
        for unit in store.get("units", []):
            ws2.cell(row=row, column=1, value=store["store_id"])
            ws2.cell(row=row, column=2, value=store["city"])
            ws2.cell(row=row, column=3, value=store["state"])
            ws2.cell(row=row, column=4, value=unit["size"])
            ws2.cell(row=row, column=5, value=unit["price"])
            ws2.cell(row=row, column=6, value="Yes" if unit["available"] else "No")
            ws2.cell(row=row, column=7, value=unit["count"])
            ws2.cell(row=row, column=8, value=unit["promo"])
            for col in range(1, 9):
                ws2.cell(row=row, column=col).font = Font(name="Calibri", size=10)
            row += 1

    p_widths = [10, 22, 8, 12, 12, 12, 12, 30]
    for col, w in enumerate(p_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(filename)
    return filename


# ── MAIN ──────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("PUBLIC STORAGE SCRAPER")
    print("=" * 60)

    # 1. Get all store IDs from sitemaps
    print("\n[1/4] Fetching state sitemaps...")
    state_urls = get_state_sitemap_urls()
    print(f"      Found {len(state_urls)} states")

    all_stores = []
    seen_ids = set()
    for i, url in enumerate(sorted(state_urls)):
        state_name = url.split("site-map-states-")[-1].replace("-", " ").title()
        print(f"      [{i+1}/{len(state_urls)}] {state_name}...", end=" ", flush=True)
        stores = get_stores_from_state(url)
        new = [s for s in stores if s["store_id"] not in seen_ids]
        for s in new:
            seen_ids.add(s["store_id"])
        all_stores.extend(new)
        print(f"{len(new)} stores")
        time.sleep(DELAY)

    print(f"\n      Total unique stores: {len(all_stores)}")

    # 2. Fetch pricing in batches
    print("\n[2/4] Fetching unit pricing from API...")
    store_ids = [s["store_id"] for s in all_stores]
    pricing_map = {}
    batches = [store_ids[i:i+BATCH_SIZE] for i in range(0, len(store_ids), BATCH_SIZE)]

    for i, batch in enumerate(batches):
        print(f"      Batch {i+1}/{len(batches)} ({len(batch)} stores)...", end=" ", flush=True)
        result = fetch_pricing_batch(batch)
        pricing_map.update(result)
        print(f"got {len(result)} responses")
        time.sleep(DELAY)

    # 3. Merge pricing into store data
    print("\n[3/4] Merging pricing data...")
    for store in all_stores:
        store["units"] = pricing_map.get(store["store_id"], [])

    # 4. Save outputs
    print("\n[4/4] Saving files...")

    # Save JSON for the website
    with open("locations.json", "w") as f:
        json.dump(all_stores, f)
    print(f"      Saved locations.json ({len(all_stores)} stores)")

    # Save Excel
    xlsx_file = save_xlsx(all_stores)
    print(f"      Saved {xlsx_file}")

    # Download both files in Colab
    print("\nDownloading files...")
    files.download("locations.json")
    files.download("public_storage_locations.xlsx")

    print("\n✅ Done! Both files downloaded.")
    print(f"   Stores scraped: {len(all_stores)}")
    priced = sum(1 for s in all_stores if s["units"])
    print(f"   Stores with pricing: {priced}")


main()
