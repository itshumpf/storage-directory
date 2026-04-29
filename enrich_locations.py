# ============================================================
# PUBLIC STORAGE ENRICHMENT SCRIPT — Run in Google Colab
# Adds site number (e.g. 77942), phone, zip, and lat/lng
# to existing locations.json
#
# HOW TO USE:
# 1. Upload your existing locations.json to Colab first
# 2. Run this script
# 3. Downloads enriched_locations.json when done
# ============================================================

# !pip install requests beautifulsoup4 openpyxl

import requests
import json
import re
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.colab import files

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
DELAY = 0.3
SAVE_EVERY = 50   # save progress every N stores in case of interruption


def extract_store_details(html, store_id):
    """
    Extract site_number, phone, zip, lat, lng from a facility page.
    The key data lives in a hidden input like:
    <input class="googleMapMarkerData" value='{"storeID":"6493","title":"77942 - Shawnee/Maurer Rd",...}' />
    """
    result = {
        "site_number": None,
        "phone": None,
        "zip": None,
        "lat": None,
        "lng": None,
    }

    soup = BeautifulSoup(html, "html.parser")

    # ── Method 1: googleMapMarkerData hidden input ──
    marker = soup.find("input", class_="googleMapMarkerData")
    if marker and marker.get("value"):
        try:
            raw = marker["value"]
            data = json.loads(raw)
            content = data.get("content", {})

            # Site number is in the title like "77942 - Shawnee/Maurer Rd"
            title = data.get("title", "") or content.get("title", "")
            m = re.match(r"^(\d{4,6})\s*-", title)
            if m:
                result["site_number"] = m.group(1)

            result["phone"] = content.get("storePhone")
            result["zip"] = content.get("postalCode")
            result["lat"] = data.get("mlat")
            result["lng"] = data.get("mlng")
        except Exception as e:
            pass

    # ── Method 2: Reviews data-sitenumber attribute ──
    if not result["site_number"]:
        el = soup.find(attrs={"data-sitenumber": True})
        if el:
            result["site_number"] = el["data-sitenumber"]

    # ── Method 3: SiteNumber in review URL ──
    if not result["site_number"]:
        review_div = soup.find(id="PLPReviews")
        if review_div:
            url_attr = review_div.get("data-ratingurl", "")
            m = re.search(r"SiteNumber=(\d+)", url_attr)
            if m:
                result["site_number"] = m.group(1)

    # ── Method 4: Phone from tel: link ──
    if not result["phone"]:
        tel = soup.find("a", href=re.compile(r"^tel:"))
        if tel:
            result["phone"] = tel["href"].replace("tel:", "").strip()

    # ── Method 5: Zip from address schema ──
    if not result["zip"]:
        zip_el = soup.find(itemprop="postalCode")
        if zip_el:
            result["zip"] = zip_el.get_text(strip=True)

    return result


def enrich_stores(stores):
    total = len(stores)
    enriched_count = 0
    already_done = sum(1 for s in stores if s.get("site_number"))

    print(f"Total stores: {total}")
    print(f"Already enriched: {already_done}")
    print(f"Remaining: {total - already_done}")
    print()

    for i, store in enumerate(stores):
        # Skip if already enriched
        if store.get("site_number"):
            continue

        url = store["url"]
        store_id = store["store_id"]

        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            resp.raise_for_status()
            details = extract_store_details(resp.text, store_id)

            store["site_number"] = details["site_number"]
            store["phone"] = details["phone"]
            store["zip"] = details["zip"]
            store["lat"] = details["lat"]
            store["lng"] = details["lng"]

            enriched_count += 1
            status = f"site#{details['site_number']}" if details["site_number"] else "no site#"
            print(f"  [{i+1}/{total}] #{store_id} {store['city']}, {store['state']} — {status} | {details['phone'] or 'no phone'}")

        except Exception as e:
            print(f"  [{i+1}/{total}] #{store_id} ERROR: {e}")
            store.setdefault("site_number", None)
            store.setdefault("phone", None)
            store.setdefault("zip", None)
            store.setdefault("lat", None)
            store.setdefault("lng", None)

        # Save progress periodically
        if enriched_count > 0 and enriched_count % SAVE_EVERY == 0:
            with open("enriched_locations.json", "w") as f:
                json.dump(stores, f)
            print(f"\n  💾 Progress saved ({enriched_count} enriched so far)\n")

        time.sleep(DELAY)

    return stores


def save_xlsx(stores, filename="public_storage_locations.xlsx"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Locations ──
    ws1 = wb.active
    ws1.title = "Locations"

    hdr_fill = PatternFill("solid", start_color="1F3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="EEF2F7")
    thin = Side(style="thin", color="CCCCCC")

    headers = ["Site #", "Store ID", "Address", "City", "State", "Zip", "Phone", "Lowest Price", "Units Available", "Lat", "Lng", "URL"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    for i, store in enumerate(stores, 2):
        units = store.get("units", [])
        avail = [u for u in units if u.get("available") and u.get("price")]
        cheapest = min(avail, key=lambda u: u["price"]) if avail else None

        vals = [
            store.get("site_number", ""),
            store["store_id"],
            store["address"],
            store["city"],
            store["state"],
            store.get("zip", ""),
            store.get("phone", ""),
            f"${cheapest['price']}/mo" if cheapest else "",
            sum(u["count"] for u in avail),
            store.get("lat", ""),
            store.get("lng", ""),
            store["url"],
        ]
        row_fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            if row_fill:
                c.fill = row_fill

    col_widths = [10, 10, 35, 20, 7, 10, 16, 14, 16, 12, 12, 65]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = w
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Unit Pricing ──
    ws2 = wb.create_sheet("Unit Pricing")
    p_headers = ["Site #", "Store ID", "City", "State", "Unit Size", "Sale Price", "Available", "Units Left", "Promotion"]
    for col, h in enumerate(p_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    row = 2
    for store in stores:
        for unit in store.get("units", []):
            ws2.cell(row=row, column=1, value=store.get("site_number", ""))
            ws2.cell(row=row, column=2, value=store["store_id"])
            ws2.cell(row=row, column=3, value=store["city"])
            ws2.cell(row=row, column=4, value=store["state"])
            ws2.cell(row=row, column=5, value=unit["size"])
            ws2.cell(row=row, column=6, value=unit["price"])
            ws2.cell(row=row, column=7, value="Yes" if unit["available"] else "No")
            ws2.cell(row=row, column=8, value=unit["count"])
            ws2.cell(row=row, column=9, value=unit.get("promo", ""))
            for col in range(1, 10):
                ws2.cell(row=row, column=col).font = Font(name="Calibri", size=10)
            row += 1

    p_widths = [10, 10, 20, 7, 12, 12, 12, 12, 30]
    for col, w in enumerate(p_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(filename)
    return filename


def main():
    print("=" * 60)
    print("PUBLIC STORAGE ENRICHMENT SCRIPT")
    print("Adds site#, phone, zip, lat/lng to locations.json")
    print("=" * 60)

    # Upload existing locations.json
    print("\nUpload your locations.json file when prompted...")
    uploaded = files.upload()
    filename = list(uploaded.keys())[0]

    with open(filename) as f:
        stores = json.load(f)

    print(f"\nLoaded {len(stores)} stores from {filename}")

    # Enrich
    print("\n[Fetching store pages for site numbers, phone, zip...]\n")
    stores = enrich_stores(stores)

    # Save final JSON
    print("\nSaving enriched_locations.json...")
    with open("enriched_locations.json", "w") as f:
        json.dump(stores, f)

    # Save Excel
    print("Saving Excel...")
    save_xlsx(stores)

    # Stats
    with_site = sum(1 for s in stores if s.get("site_number"))
    with_phone = sum(1 for s in stores if s.get("phone"))
    print(f"\n✅ Done!")
    print(f"   Stores with site number: {with_site}/{len(stores)}")
    print(f"   Stores with phone:       {with_phone}/{len(stores)}")

    # Download
    files.download("enriched_locations.json")
    files.download("public_storage_locations.xlsx")


main()
